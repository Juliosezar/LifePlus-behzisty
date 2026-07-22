import json
import logging
import re
import datetime
import openpyxl
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin

from .models import Conversation, ChatMessage
from .agent import ask_question

logger = logging.getLogger(__name__)


def _serialize_result(rows):
    """Convert non-JSON-serializable types (datetime, date, Decimal) to strings."""
    if not rows or not isinstance(rows, list):
        return rows
    clean = []
    for row in rows:
        if isinstance(row, dict):
            clean.append({
                k: v.isoformat() if isinstance(v, (datetime.datetime, datetime.date)) else str(v) if v is not None and not isinstance(v, (int, float, bool, str)) else v
                for k, v in row.items()
            })
        elif isinstance(row, (list, tuple)):
            clean.append([
                v.isoformat() if isinstance(v, (datetime.datetime, datetime.date)) else str(v) if v is not None and not isinstance(v, (int, float, bool, str)) else v
                for v in row
            ])
        else:
            clean.append(row)
    return clean


def _parse_markdown_table(text):
    """Extract a markdown table from text and return (headers, rows) or (None, None)."""
    lines = text.strip().split('\n')
    table_lines = []
    in_table = False
    for line in lines:
        stripped = line.strip()
        if stripped.startswith('|') and stripped.endswith('|') and stripped.count('|') >= 3:
            in_table = True
            table_lines.append(stripped)
        elif in_table and stripped == '':
            break
        elif in_table:
            break

    if len(table_lines) < 2:
        return None, None

    # First line = headers
    headers = [h.strip() for h in table_lines[0].split('|')[1:-1]]
    # Second line = separator (skip)
    # Rest = data rows
    rows = []
    for line in table_lines[2:]:
        cells = [c.strip() for c in line.split('|')[1:-1]]
        if len(cells) == len(headers):
            rows.append(dict(zip(headers, cells)))

    return headers if rows else None, rows


def _strip_markdown_tables(text):
    """Remove markdown tables from LLM reply text."""
    lines = text.split('\n')
    result = []
    in_table = False
    for line in lines:
        stripped = line.strip()
        # Detect table rows (start with | and contain |)
        if stripped.startswith('|') and stripped.endswith('|') and stripped.count('|') >= 3:
            in_table = True
            continue
        # Detect table separator lines (|---|---|)
        if re.match(r'^[\|\s\-:]+$', stripped) and '|' in stripped:
            in_table = True
            continue
        if in_table and stripped == '':
            in_table = False
            continue
        if not in_table:
            result.append(line)
    return '\n'.join(result).strip()
    """Convert non-JSON-serializable types (datetime, date, Decimal) to strings."""
    if not rows or not isinstance(rows, list):
        return rows
    clean = []
    for row in rows:
        if isinstance(row, dict):
            clean.append({
                k: v.isoformat() if isinstance(v, (datetime.datetime, datetime.date)) else str(v) if v is not None and not isinstance(v, (int, float, bool, str)) else v
                for k, v in row.items()
            })
        elif isinstance(row, (list, tuple)):
            clean.append([
                v.isoformat() if isinstance(v, (datetime.datetime, datetime.date)) else str(v) if v is not None and not isinstance(v, (int, float, bool, str)) else v
                for v in row
            ])
        else:
            clean.append(row)
    return clean


class ChatPageView(LoginRequiredMixin, TemplateView):
    template_name = 'ai_chat/chat.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['conversations'] = Conversation.objects.filter(
            user=self.request.user
        )[:20]

        conv_id = self.request.GET.get('conv')
        if conv_id:
            try:
                context['active_conversation'] = Conversation.objects.get(
                    id=conv_id, user=self.request.user
                )
                context['messages'] = context['active_conversation'].messages.all()
            except Conversation.DoesNotExist:
                pass

        return context


class ChatAPIView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'error': 'Invalid JSON'}, status=400)

        user_message = data.get('message', '').strip()
        conversation_id = data.get('conversation_id')

        if not user_message:
            return JsonResponse({'status': 'error', 'error': 'Empty message'}, status=400)

        # Get or create conversation
        if conversation_id:
            try:
                conversation = Conversation.objects.get(
                    id=conversation_id, user=request.user
                )
            except Conversation.DoesNotExist:
                return JsonResponse({'status': 'error', 'error': 'Conversation not found'}, status=404)
        else:
            conversation = Conversation.objects.create(
                user=request.user,
                title=user_message[:60],
            )

        # Save user message
        ChatMessage.objects.create(
            conversation=conversation,
            role='user',
            content=user_message,
        )

        # Build context: last 6 messages for continuity
        recent_messages = conversation.messages.order_by('-created_at')[:6]
        history_text = ""
        for msg in reversed(list(recent_messages)):
            role_label = "کاربر" if msg.role == 'user' else "دستیار"
            history_text += f"{role_label}: {msg.content}\n"

        # Ask the LLM + execute SQL
        error_text = ""
        try:
            response = ask_question(user_message, history=history_text)
            reply_text = response['reply']
            sql_query = response['sql_query']
            sql_result = response['result']
            columns = response['columns']
        except Exception as e:
            logger.exception("AI chat error")
            reply_text = "متأسفانه در پردازش سوال شما خطایی رخ داد. لطفاً دوباره تلاش کنید."
            sql_query = ""
            sql_result = []
            columns = []
            error_text = str(e)

        # Save assistant message
        assistant_msg = ChatMessage.objects.create(
            conversation=conversation,
            role='assistant',
            content=reply_text,
            sql_query=sql_query,
            sql_result=_serialize_result(sql_result) if sql_result else None,
            error=error_text,
        )

        # Auto-title on first exchange
        if conversation.messages.count() <= 2 and conversation.title == user_message[:60]:
            conversation.title = user_message[:80]
            conversation.save(update_fields=['title'])

        return JsonResponse({
            'status': 'ok',
            'conversation_id': conversation.id,
            'message_id': assistant_msg.id,
            'reply': reply_text,
            'sql_query': sql_query,
            'sql_result': sql_result,
            'columns': columns,
            'error': error_text,
        })


class ExportExcelView(LoginRequiredMixin, View):
    def get(self, request, message_id, *args, **kwargs):
        try:
            msg = ChatMessage.objects.get(
                id=message_id,
                conversation__user=request.user,
                role='assistant',
            )
        except ChatMessage.DoesNotExist:
            return JsonResponse({'error': 'Not found'}, status=404)

        if not msg.sql_result or not isinstance(msg.sql_result, list):
            return JsonResponse({'error': 'No data to export'}, status=400)

        rows = msg.sql_result
        columns = list(rows[0].keys()) if rows else []

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "نتایج پرسش"
        ws.sheet_view.rightToLeft = True

        # Headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(row.get(col_name, '')))

        # Auto-width columns
        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(str(col_name))
            for row in rows[:100]:
                val = str(row.get(col_name, ''))
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="ai_chat_result.xlsx"'
        wb.save(response)
        return response
